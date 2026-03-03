# Core views - handles dashboard logic
# Shows different stats based on user role (admin/teacher/student)

"""
core/views.py — Dashboard view and data export views (Excel and PDF).

Provides:
  dashboard     — role-specific summary page shown after login
  export_excel  — download all CMS data as a styled .xlsx file (admin only)
  export_pdf    — download all CMS data as a formatted PDF file (admin only)
"""

import io  # Used to build in-memory file buffers for streaming downloads

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse

# Import all models needed for dashboard stats and exports
from users.models import CustomUser
from courses.models import Course
from classes.models import Classroom
from attendance.models import AttendanceSession, AttendanceRecord
from assignments.models import Assignment, AssignmentSubmission


# ─── Dashboard ────────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    """
    Render the main dashboard with role-specific stats and recent activity.

    Admin   → system-wide counts and recent sessions / assignments
    Teacher → their own classes, student count, recent activity
    Student → their enrolled courses/classes, recent assignments, attendance %
    """
    u   = request.user
    ctx = {}  # Context dictionary passed to the template

    if u.is_admin:
        # ── Admin stats: system-wide counts ───────────────────────────────────
        ctx['total_students'] = CustomUser.objects.filter(role='student').count()
        ctx['total_teachers'] = CustomUser.objects.filter(role='teacher').count()
        ctx['total_courses']  = Course.objects.count()
        ctx['total_classes']  = Classroom.objects.count()

        # Recent activity: last 5 attendance sessions and assignments
        ctx['recent_sessions'] = AttendanceSession.objects.select_related(
            'classroom', 'taken_by'
        ).order_by('-date')[:5]

        ctx['recent_assignments'] = Assignment.objects.select_related(
            'classroom', 'uploaded_by'
        ).order_by('-created_at')[:5]

    elif u.is_teacher:
        # ── Teacher stats: scoped to their own classes ─────────────────────────
        my_classes = Classroom.objects.filter(teacher=u)
        ctx['my_classes'] = my_classes

        # Count total students across all their classrooms
        ctx['total_students'] = sum(c.students.count() for c in my_classes)

        # Number of courses they teach
        ctx['my_courses'] = Course.objects.filter(teacher=u).count()

        # Recent assignments they posted and sessions they took
        ctx['recent_assignments'] = Assignment.objects.filter(
            uploaded_by=u
        ).order_by('-created_at')[:5]

        ctx['recent_sessions'] = AttendanceSession.objects.filter(
            taken_by=u
        ).order_by('-date')[:5]

    else:
        # ── Student stats: scoped to their enrollment ─────────────────────────
        ctx['my_courses']  = u.enrolled_courses.all()
        ctx['my_classes']  = u.enrolled_classes.select_related('teacher', 'course').all()

        # Assignments posted to any classroom this student is enrolled in
        ctx['recent_assignments'] = Assignment.objects.filter(
            classroom__in=u.enrolled_classes.all()
        ).order_by('-created_at')[:5]

        # Overall attendance percentage
        ctx['attendance_pct'] = _student_attendance_pct(u)

    return render(request, 'core/dashboard.html', ctx)


def _student_attendance_pct(student):
    """
    Calculate the overall attendance percentage for a student.

    Counts all their AttendanceRecords, then finds how many were
    'present' or 'late' (both count as attended).

    Returns 0 if the student has no records yet (avoids division by zero).
    """
    # Total number of attendance records for this student
    total = AttendanceRecord.objects.filter(student=student).count()

    # Records where they were marked present or late (both count as "attended")
    present = AttendanceRecord.objects.filter(
        student=student, status__in=['present', 'late']
    ).count()

    # Avoid division by zero; return 0% if no records exist yet
    return round(present / total * 100, 1) if total > 0 else 0


# ─── Excel Export ─────────────────────────────────────────────────────────────

@login_required
def export_excel(request):
    """
    Export CMS data to a styled multi-sheet Excel (.xlsx) file (admin only).

    Sheets produced:
      1. Summary   — system-wide counts
      2. Students  — full student roster
      3. Teachers  — full teacher roster
      4. Courses   — course list
      5. Assignments — assignment list with submission counts

    Requires openpyxl (pip install openpyxl).
    Streams the file directly to the browser as a download response.
    """
    # Restrict to admin users
    if not request.user.is_admin:
        return redirect('dashboard')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        # Inform the admin if openpyxl is missing rather than crashing
        return HttpResponse(
            "openpyxl is not installed. Run: python -m pip install openpyxl",
            status=500,
        )

    # Create a new workbook
    wb = openpyxl.Workbook()

    # Style definitions for header rows: bold white text on indigo background
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4F46E5")  # Indigo

    def style_headers(ws):
        """Apply bold/indigo style to the first (header) row of a worksheet."""
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active           # The default sheet created with the workbook
    ws1.title = "Summary"
    ws1.append(["Metric", "Count"])   # Header row
    style_headers(ws1)

    # Append one row per metric
    ws1.append(["Total Students",    CustomUser.objects.filter(role='student').count()])
    ws1.append(["Total Teachers",    CustomUser.objects.filter(role='teacher').count()])
    ws1.append(["Total Admins",      CustomUser.objects.filter(role='admin').count()])
    ws1.append(["Total Courses",     Course.objects.count()])
    ws1.append(["Total Classes",     Classroom.objects.count()])
    ws1.append(["Total Assignments", Assignment.objects.count()])
    ws1.append(["Total Submissions", AssignmentSubmission.objects.count()])

    # Set readable column widths
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 15

    # ── Sheet 2: Students ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Students")
    ws2.append(["#", "Full Name", "Username", "Email", "Phone", "Date of Birth"])
    style_headers(ws2)

    # Enumerate students ordered by last name (idx starts at 1 for row numbers)
    for idx, s in enumerate(CustomUser.objects.filter(role='student').order_by('last_name'), 1):
        ws2.append([
            idx,
            s.get_full_name() or s.username,  # Fall back to username if no full name
            s.username,
            s.email or "",
            str(s.phone or ""),
            str(s.date_of_birth or ""),
        ])

    for col in ["A", "B", "C", "D", "E", "F"]:
        ws2.column_dimensions[col].width = 20

    # ── Sheet 3: Teachers ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Teachers")
    ws3.append(["#", "Full Name", "Username", "Email", "Phone"])
    style_headers(ws3)

    for idx, t in enumerate(CustomUser.objects.filter(role='teacher').order_by('last_name'), 1):
        ws3.append([
            idx,
            t.get_full_name() or t.username,
            t.username,
            t.email or "",
            str(t.phone or ""),
        ])

    for col in ["A", "B", "C", "D", "E"]:
        ws3.column_dimensions[col].width = 22

    # ── Sheet 4: Courses ──────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Courses")
    ws4.append(["#", "Course Name", "Course Code"])
    style_headers(ws4)

    for idx, c in enumerate(Course.objects.all(), 1):
        # getattr with fallback handles models that may use different field names
        code = getattr(c, 'code', getattr(c, 'course_code', ''))
        ws4.append([idx, c.name, str(code)])

    for col in ["A", "B", "C"]:
        ws4.column_dimensions[col].width = 25

    # ── Sheet 5: Assignments ──────────────────────────────────────────────────
    ws5 = wb.create_sheet("Assignments")
    ws5.append(["#", "Title", "Classroom", "Uploaded By", "Due Date", "Max Marks", "Submissions"])
    style_headers(ws5)

    for idx, a in enumerate(
        Assignment.objects.select_related('classroom', 'uploaded_by').all(), 1
    ):
        ws5.append([
            idx,
            a.title,
            str(a.classroom),
            a.uploaded_by.get_full_name() or a.uploaded_by.username,
            str(a.due_date or "No deadline"),
            a.max_marks,
            a.submissions.count(),  # Count related submissions (extra query per row)
        ])

    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws5.column_dimensions[col].width = 22

    # ── Stream the workbook as a download ─────────────────────────────────────
    buffer = io.BytesIO()       # In-memory buffer — no temp file needed
    wb.save(buffer)             # Write workbook bytes into the buffer
    buffer.seek(0)              # Rewind buffer to the beginning for reading

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    # Tell the browser to download the file with this suggested name
    response['Content-Disposition'] = 'attachment; filename="cms_report.xlsx"'
    return response


# ─── PDF Export ───────────────────────────────────────────────────────────────

@login_required
def export_pdf(request):
    """
    Export CMS data to a formatted PDF file (admin only).

    Uses ReportLab to build a document with:
      - A title / header paragraph
      - A summary table
      - A students table
      - A teachers table
      - A courses table

    Requires reportlab (pip install reportlab).
    Streams the PDF directly to the browser as a download response.
    """
    if not request.user.is_admin:
        return redirect('dashboard')

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        return HttpResponse(
            "reportlab is not installed. Run: python -m pip install reportlab",
            status=500,
        )

    # Build the PDF into an in-memory buffer
    buffer = io.BytesIO()

    # SimpleDocTemplate handles page layout, margins, and page breaks
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm,  bottomMargin=2*cm,
    )

    # Load default paragraph styles and define our custom brand color
    styles  = getSampleStyleSheet()
    primary = colors.HexColor('#4F46E5')  # Indigo — matches the CSS theme
    elements = []  # List of ReportLab flowables (paragraphs, tables, spacers)

    # ── Custom paragraph styles ────────────────────────────────────────────────
    title_style = ParagraphStyle(
        'CMSTitle',
        parent=styles['Title'],
        textColor=primary,
        fontSize=20,
        spaceAfter=6,
    )
    heading_style = ParagraphStyle(
        'CMSHeading',
        parent=styles['Heading2'],
        textColor=primary,
        spaceBefore=16,
        spaceAfter=6,
    )

    # ── Document header ────────────────────────────────────────────────────────
    elements.append(Paragraph("College Management System — Report", title_style))
    elements.append(Paragraph(
        "Generated by: " + (request.user.get_full_name() or request.user.username),
        styles['Normal'],
    ))
    elements.append(Spacer(1, 0.4*cm))  # Vertical whitespace

    def make_table(data, col_widths=None):
        """
        Build a styled ReportLab Table from a list of row lists.

        data       — list of lists; first row is treated as the header
        col_widths — optional list of column widths (in ReportLab units)

        Styling applied:
          - Header row: indigo background, white bold text
          - Data rows: alternating white / light-grey background
          - All cells: light grey grid border, vertically centred
        """
        t = Table(data, colWidths=col_widths, repeatRows=1)  # repeatRows: repeat header on each page
        t.setStyle(TableStyle([
            # Header row formatting
            ('BACKGROUND',    (0, 0), (-1, 0),  primary),
            ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
            ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, 0),  10),
            ('TOPPADDING',    (0, 0), (-1, 0),  8),
            ('BOTTOMPADDING', (0, 0), (-1, 0),  8),
            # Data rows: alternate white and light grey for readability
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F1F5F9')]),
            ('FONTSIZE',      (0, 1), (-1, -1),  9),
            ('TOPPADDING',    (0, 1), (-1, -1),  6),
            ('BOTTOMPADDING', (0, 1), (-1, -1),  6),
            # Light grid lines between all cells
            ('GRID',   (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            # Vertically centre all cell content
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        return t

    # ── Summary table ──────────────────────────────────────────────────────────
    elements.append(Paragraph("Summary", heading_style))
    elements.append(make_table([
        ["Metric", "Count"],
        ["Total Students",    CustomUser.objects.filter(role='student').count()],
        ["Total Teachers",    CustomUser.objects.filter(role='teacher').count()],
        ["Total Admins",      CustomUser.objects.filter(role='admin').count()],
        ["Total Courses",     Course.objects.count()],
        ["Total Classes",     Classroom.objects.count()],
        ["Total Assignments", Assignment.objects.count()],
        ["Total Submissions", AssignmentSubmission.objects.count()],
    ], col_widths=[10*cm, 5*cm]))

    # ── Students table ─────────────────────────────────────────────────────────
    elements.append(Spacer(1, 0.4*cm))
    elements.append(Paragraph("Students", heading_style))

    rows = [["#", "Full Name", "Username", "Email"]]
    for idx, s in enumerate(CustomUser.objects.filter(role='student').order_by('last_name'), 1):
        rows.append([str(idx), s.get_full_name() or s.username, s.username, s.email or "—"])

    # Show the table only if there is at least one data row; otherwise show a message
    elements.append(
        make_table(rows, col_widths=[1*cm, 5*cm, 4*cm, 6*cm])
        if len(rows) > 1
        else Paragraph("No students found.", styles['Normal'])
    )

    # ── Teachers table ─────────────────────────────────────────────────────────
    elements.append(Spacer(1, 0.4*cm))
    elements.append(Paragraph("Teachers", heading_style))

    rows = [["#", "Full Name", "Username", "Email"]]
    for idx, t in enumerate(CustomUser.objects.filter(role='teacher').order_by('last_name'), 1):
        rows.append([str(idx), t.get_full_name() or t.username, t.username, t.email or "—"])

    elements.append(
        make_table(rows, col_widths=[1*cm, 5*cm, 4*cm, 6*cm])
        if len(rows) > 1
        else Paragraph("No teachers found.", styles['Normal'])
    )

    # ── Courses table ──────────────────────────────────────────────────────────
    elements.append(Spacer(1, 0.4*cm))
    elements.append(Paragraph("Courses", heading_style))

    rows = [["#", "Course Name", "Code"]]
    for idx, c in enumerate(Course.objects.all(), 1):
        code = getattr(c, 'code', getattr(c, 'course_code', ''))
        rows.append([str(idx), c.name, str(code)])

    elements.append(
        make_table(rows, col_widths=[1*cm, 9*cm, 6*cm])
        if len(rows) > 1
        else Paragraph("No courses found.", styles['Normal'])
    )

    # ── Build and stream the PDF ───────────────────────────────────────────────
    doc.build(elements)   # ReportLab assembles all flowables into pages
    buffer.seek(0)        # Rewind buffer before reading

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="cms_report.pdf"'
    return response
